from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
import cv2
import datetime
import os

resolution = 20 #decrease to increase (divides original h and w)
rowH = 25 #actual height of the rows
colW = 3 #actual width of the columns
zoom = 60 #zoom factor (so you can see the entire image when opening the worksheet)

def initSheet(resW, resH): #creates the sheet
    for i in range(1,resH+1):
        for j in range(1,resW+1):
            ws.cell(row=i,column=j)

def toHex(r,g,b): #convert rgb to hex
    hexchars = "0123456789ABCDEF"
    finalColor = hexchars[r // 16] + hexchars[r % 16] + hexchars[g // 16] + hexchars[g % 16] + hexchars[b // 16] + hexchars[b % 16]
    return finalColor

def setCellColor(r,g,b):
    color = toHex(r,g,b)
    cellColor = PatternFill(fill_type='solid', start_color = color)
    return cellColor

def resizeImage(imageName, resH,resW):#resize the image - lower the resolution for translation to excel cells
    image = cv2.imread(imageName)
    smallImage = cv2.resize(image,(resW,resH))
    height, width, channel = smallImage.shape
    return smallImage

def setCellDims(worksheet, resH, resW): #adjust row/column height/width
    for i in range(1, resW+1):
        ws.row_dimensions[i].height = rowH
    for i in range(1, resH+1):
        ws.column_dimensions[get_column_letter(i)].width = colW

def img2excel(): #actually get the image to Excel
    initSheet(resW, resH)
    pixelList = resizeImage(imageName, resH, resW)
    i=0
    j=0
    for row in ws.iter_rows(min_col=1, max_col = resW, max_row = resH):
        j=0
        for cell in row: #uses bgr not rgb because of opencv2
            b = pixelList[i][j][0]
            g = pixelList[i][j][1]
            r = pixelList[i][j][2]
            fill = setCellColor(r,g,b)
            cell.fill = fill
            j+=1
        i+=1

def camScreenCap():
    outputName = "capture_"+str(datetime.datetime.now().time())[:8]+".png"
    camera = cv2.VideoCapture(0)
    ret, frame = camera.read()
    cv2.imwrite(outputName,frame)
    camera.release()
    cv2.destroyAllWindows()
    return outputName
#Main
wb = Workbook() #Initialize workbook and sheet in memory
ws = wb.worksheets[0]
ws.title='Excelfie'

imageName = camScreenCap()
img = cv2.imread(imageName)
h,w,c = img.shape #get height and width (c is also returned by .shape but not used)
resW = int(w/resolution)
resH = int(h/resolution)
img2excel()

setCellDims(ws,resW,resH)
ws.sheet_view.zoomScale = zoom #set zoom so that it looks right when you open wb
wb.save('excelfie.xlsx')
os.remove(imageName) #delete screenshot file, leaving only the workbook
