import xlwings as xw
import cv2
import datetime
import openpyxl
from openpyxl import Workbook
# from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule
import numpy as np
import os

resolution = 20 #decrease to increase (divides original h and w)
rowH = 25 #actual height of the rows
colW = 1 #actual width of the columns
zoom = 60 #zoom factor (so you can see the entire image when opening the worksheet)

def initSheet(resW, resH): # Set up the initial sheet with the right number of rows and columns
    for i in range(1,resH+1):
        for j in range(1,resW+1):
            ws.cell(row=i,column=j)

def resizeImage(imageName, resH,resW):#lower the resolution for translation to excel cells
    # image = frame
    image = cv2.imread(imageName)
    smallImage = cv2.resize(image,(resW,resH))
    height, width, channel = smallImage.shape
    return smallImage

def setCellDims(worksheet, resH, resW): #adjust row/column height/width
    for i in range(1, resW+1):
        ws.row_dimensions[i].height = rowH
    for i in range(1, resH+1):
        ws.column_dimensions[get_column_letter(i)].width = colW

def setCondFormatting(worksheet,resH,resW): #bgr not rgb because of opencv2
    redRule = ColorScaleRule(start_type='num', start_value=0,
    start_color='00000000',end_type='num', end_value=255,
    end_color='00FF0000')

    greenRule = ColorScaleRule(start_type='num', start_value=0,
    start_color='00000000',end_type='num', end_value=255,
    end_color='0000FF00')

    blueRule = ColorScaleRule(start_type='num', start_value=0,
    start_color='00000000',end_type='num', end_value=255,
    end_color='000000FF')

    for i in range(1,resW*3+1): #assign b, g, r condformatting to every 3 columns
        rng = f'{get_column_letter(i)}1:{get_column_letter(i)}{resW}'
        if (i+2)%3 == 0:
            ws.conditional_formatting.add(rng,blueRule)
        elif (i+1)%3 == 0:
            ws.conditional_formatting.add(rng,greenRule)
        else:
            ws.conditional_formatting.add(rng,redRule)

def camScreenCap(): #take a screenshot, return the name
    outputName = "capture_"+str(datetime.datetime.now().time())[:8]+".png"
    camera = cv2.VideoCapture(0)
    ret, frame = camera.read()
    cv2.imwrite(outputName,frame)
    camera.release()
    cv2.destroyAllWindows()
    return outputName
#Main
bookName = 'excelfieVid.xlsx' #Create workbook + sheet in memory using openpyxl
wb = Workbook()
ws = wb.worksheets[0]
ws.title='Excelfie'

imageName = camScreenCap() #Take initial screenshot from which to base sizes
img = cv2.imread(imageName)
h,w,c = img.shape #get height and width (c is also returned by .shape but not used)
resW = int(w/resolution)*3 # *3 because it's 3 cells/pixel (for b, g, r)
resW3 = resW//3 #used for resizing the image
resH = int(h/resolution)
initSheet(resW,resH)
pixelList = resizeImage(imageName, resH, resW//3)

setCellDims(ws,resW,resH)
ws.sheet_view.zoomScale = zoom #set zoom so that it looks right when you open wb
setCondFormatting(ws,resH,resW)
wb.save(bookName)
os.remove(imageName) #delete screenshot used for sizing

wb = xw.Book(bookName) #set up wb and sheet in xlwings
ws = wb.sheets[0]

camera = cv2.VideoCapture(0) #start camera

while True:
    ret, frame = camera.read() #get image info - frame contains pixel info
    pixelList = cv2.resize(frame, (resW3, resH))
    pixelList = np.reshape(np.ravel(pixelList),(resH,resW)).tolist() #reshape the array so it can be output to cells
    ws.range((1,1)).value = pixelList
